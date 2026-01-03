from openpyxl import load_workbook, Workbook

def read_headers(path: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    headers = []
    for cell in ws[1]:
        headers.append("" if cell.value is None else str(cell.value).strip())
    wb.close()
    return [h for h in headers if h != ""]

def read_rows_as_dicts(path: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    header_row = [("" if c.value is None else str(c.value).strip()) for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, h in enumerate(header_row):
            if h == "":
                continue
            v = r[i] if i < len(r) else None
            row_dict[h] = v
        rows.append(row_dict)
    wb.close()
    return header_row, rows

def write_failed_rows(out_path: str, header_row: list, failed_rows: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.append(header_row)
    for row in failed_rows:
        ws.append([row.get(h, None) for h in header_row])
    wb.save(out_path)
